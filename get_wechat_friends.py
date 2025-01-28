#用于爬取微信好友列表
import datetime
import uiautomation as uia
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import threading
import queue
import time
import pythoncom

class WeChatContactsCrawler:
    def __init__(self):
        self.contacts_queue = queue.Queue()
        self.is_crawling = True
        self.current_time = datetime.datetime.now().strftime("%Y-%m-%d")
        self.filename = f"list.xlsx"
        self.headers = ["code", "nickname", "area", "remark", "tag", "sign", "from"]
        self.headers_name = ["微信号", "昵称", "地区", "备注", "标签", "签名", "来源"]

    def init_excel(self):
        """初始化Excel文件"""
        wb = Workbook()
        ws = wb.active
        # 设置标题样式
        font_color = Font(color="FFFFFF")
        fill_color = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        
        for col_num, header in enumerate(self.headers_name, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = font_color
            cell.fill = fill_color
        wb.save(self.filename)

    def init_wechat_window(self):
        """初始化微信窗口"""
        wechat_window = uia.WindowControl(ClassName='WeChatMainWndForPC')
        wechat_window.SwitchToThisWindow()
        wechat_window.MoveToCenter()
        return wechat_window

    def scroll_to_top(self, wechat_window):
        """滚动到联系人列表顶部"""
        prevTop = ""
        sameTopCount = 0
        while sameTopCount < 2:
            session_list = wechat_window.ListControl(Name='联系人')
            currentTop = session_list.GetChildren()[0].Name
            if currentTop == prevTop:
                sameTopCount += 1
            else:
                sameTopCount = 0
            prevTop = currentTop
            session_list.WheelUp(wheelTimes=20, waitTime=0.1)

    def crawl_contacts(self):
        """爬取联系人信息的线程函数"""
        # 在线程中初始化COM
        pythoncom.CoInitialize()
        try:
            wechat_window = self.init_wechat_window()
            toolBar = wechat_window.ToolBarControl(Name="导航")
            toolBar.GetChildren()[2].Click()  # 点击通讯录按钮
            
            # 滚动到顶部
            self.scroll_to_top(wechat_window)
            
            # 获取并点击第一个联系人
            session_list = wechat_window.ListControl(Name='联系人').GetChildren()
            for index, item in reversed(list(enumerate(session_list))):
                if item.Name == "":
                    break
            session_list[index + 1].Click()

            preWechatCode = ""
            while self.is_crawling:
                try:
                    wechatCodeTag = wechat_window.TextControl(Name="微信号：")
                    if not wechatCodeTag.Exists(0.1):
                        wechat_window.SendKeys("{DOWN}")
                        continue

                    contact = {
                        "code": wechatCodeTag.GetNextSiblingControl().Name,
                        "nickname": "",
                        "area": "",
                        "remark": "",
                        "tag": "",
                        "sign": "",
                        "from": ""
                    }

                    if preWechatCode == contact["code"]:
                        self.is_crawling = False
                        break

                    preWechatCode = contact["code"]
                    self.get_contact_details(wechat_window, contact)
                    self.contacts_queue.put(contact)
                    wechat_window.SendKeys("{DOWN}")

                except Exception as e:
                    print(f"获取联系人信息出错: {str(e)}")
                    continue

        except Exception as e:
            print(f"爬虫线程异常: {str(e)}")
        finally:
            self.contacts_queue.put(None)  # 发送结束信号
            pythoncom.CoUninitialize()  # 清理COM

    def get_contact_details(self, wechat_window, contact):
        """获取联系人详细信息"""
        contact["nickname"] = wechat_window.ButtonControl(Name="更多").GetPreviousSiblingControl().Name
        
        nicknameTag = wechat_window.TextControl(Name="昵称：")
        if nicknameTag.Exists(0.1):
            contact["remark"] = contact["nickname"]
            contact["nickname"] = nicknameTag.GetNextSiblingControl().Name

        for field, label in [
            ("area", "地区："),
            ("sign", "个性签名"),
            ("tag", "标签"),
            ("from", "来源")
        ]:
            tag = wechat_window.TextControl(Name=label)
            if tag.Exists(0.1):
                contact[field] = tag.GetNextSiblingControl().Name

    def save_to_excel(self):
        """保存数据到Excel的线程函数"""
        # Excel操作也需要COM初始化
        pythoncom.CoInitialize()
        try:
            while True:
                contact = self.contacts_queue.get()
                if contact is None:  # 结束信号
                    break
                    
                try:
                    wb = load_workbook(self.filename)
                    ws = wb.active
                    row = [contact[key] for key in self.headers]
                    ws.append(row)
                    wb.save(self.filename)
                    print(f"已保存联系人: {contact['nickname']}")
                except Exception as e:
                    print(f"保存到Excel出错: {str(e)}")
                finally:
                    self.contacts_queue.task_done()
        finally:
            pythoncom.CoUninitialize()  # 清理COM

def main():
    # 主线程初始化COM
    pythoncom.CoInitialize()
    try:
        crawler = WeChatContactsCrawler()
        crawler.init_excel()
        
        # 创建并启动线程
        crawl_thread = threading.Thread(target=crawler.crawl_contacts)
        save_thread = threading.Thread(target=crawler.save_to_excel)
        
        crawl_thread.start()
        save_thread.start()
        
        # 等待线程结束
        crawl_thread.join()
        save_thread.join()
        
        print("通讯录获取完成！")
    finally:
        pythoncom.CoUninitialize()

if __name__ == '__main__':
    main()
